import io
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from PIL import Image as PILImage, ImageDraw, ImageFont
from bson import ObjectId
import bson

from utils.db import get_mongo_db
from utils.helpers import get_workdays

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#64748b"))
        
        # Header
        self.drawString(54, 750, "EM Sprint Cockpit - Sprint Report")
        self.setStrokeColor(colors.HexColor("#e2e8f0"))
        self.setLineWidth(0.5)
        self.line(54, 742, 558, 742)
        
        # Footer
        self.line(54, 45, 558, 45)
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(558, 30, page_text)
        self.drawString(54, 30, "Confidential - For Internal Management Review Only")
        self.restoreState()

def generate_sprint_report_data(sprint_id):
    """Gathers and calculates all stats for the given sprint_id."""
    db = get_mongo_db()
    sprint = db['sprints'].find_one({"_id": ObjectId(sprint_id)})
    if not sprint:
        raise ValueError("Sprint not found.")

    team_id = sprint['team_id']
    team_doc = db['teams'].find_one({"_id": ObjectId(team_id)})
    team_name = team_doc['name'] if team_doc else "Unknown Team"

    # 1. Roster and capacity setup
    users_cursor = db['users'].find({"team_id": str(team_id), "user_role": {"$ne": "Super Admin"}})
    team_list = list(users_cursor)
    
    s_start = sprint['start_date']
    s_end = sprint['end_date']
    
    # Read leaves & holidays
    leaves_cursor = db['leaves'].find({
        "team_id": str(team_id),
        "$or": [{"sprint_id": str(sprint_id)}, {"sprint_id": "0"}, {"sprint_id": 0}]
    })
    leaves_list = list(leaves_cursor)
    leaves_df = pd.DataFrame(leaves_list) if leaves_list else pd.DataFrame()
    
    hols_cursor = db['holidays'].find({
        "team_id": str(team_id),
        "sprint_id": str(sprint_id),
        "holiday_date": {"$gte": str(s_start), "$lte": str(s_end)}
    })
    hols_list = list(hols_cursor)
    holiday_count = len(hols_list)
    work_days = get_workdays(s_start, s_end)

    # Individual capacities
    dev_details = []
    total_net_sp = 0.0
    bug_buffer_allocated = 0.0
    adhoc_buffer_allocated = 0.0
    ceremony_buffer_allocated = 0.0
    plannable_capacity = 0.0

    # Backlog tasks
    backlog_cursor = db['backlog'].find({"team_id": str(team_id), "sprint_id": str(sprint_id)})
    tasks_list = list(backlog_cursor)
    tasks_df = pd.DataFrame(tasks_list) if tasks_list else pd.DataFrame()

    for dev in team_list:
        dev_name = dev['name']
        l_days = 0
        if not leaves_df.empty:
            d_leaves = leaves_df[leaves_df['name'] == dev_name]
            for _, l in d_leaves.iterrows():
                l_s = max(pd.to_datetime(l['start_date']).date(), pd.to_datetime(s_start).date())
                l_e = min(pd.to_datetime(l['end_date']).date(), pd.to_datetime(s_end).date())
                if l_s <= l_e:
                    l_days += get_workdays(l_s, l_e)

        dev_role = dev.get('role', '')
        daily_sp = 0.0 if dev_role in ['PM', 'EM'] else dev.get('daily_sp', 0.0)
        net_capacity = max(work_days - l_days - holiday_count, 0) * daily_sp
        dev_eff_days = max(work_days - l_days - holiday_count, 0)
        
        b_p = dev.get('bug_p', 0.0)
        a_p = dev.get('adhoc_p', 0.0)
        c_p = dev.get('ceremony_p', 0.0)

        dev_bug = b_p
        dev_adhoc = a_p
        dev_cere = c_p
        dev_plan = net_capacity - (dev_bug + dev_adhoc + dev_cere)

        # Developer backlog
        allocated = 0.0
        completed = 0.0
        if not tasks_df.empty:
            dev_tasks = tasks_df[tasks_df['assignee'] == dev_name]
            allocated = dev_tasks['sp'].sum() if not dev_tasks.empty else 0.0
            completed = dev_tasks[dev_tasks['status'] == 'Done']['sp'].sum() if not dev_tasks.empty else 0.0

        remaining = max(net_capacity - allocated, 0.0)

        dev_details.append({
            "name": dev_name,
            "role": dev.get('role', 'Developer'),
            "capacity": net_capacity,
            "allocated": allocated,
            "completed": completed,
            "remaining": remaining,
            "eff_days": dev_eff_days,
            "leaves": l_days,
            "daily_sp": daily_sp,
            "delivery_rate": (completed / allocated * 100.0) if allocated > 0.0 else 0.0,
            "utilization_rate": (allocated / net_capacity * 100.0) if net_capacity > 0 else 0.0,
        })

        total_net_sp += net_capacity
        bug_buffer_allocated += dev_bug
        adhoc_buffer_allocated += dev_adhoc
        ceremony_buffer_allocated += dev_cere
        plannable_capacity += dev_plan

    # Backlog statistics
    planned_sp = 0.0
    delivered_sp = 0.0
    planned_tickets = 0
    completed_tickets = 0
    spillover_tickets = 0

    bug_tickets_planned = 0
    bug_tickets_done = 0
    bug_buffer_used = 0.0
    adhoc_buffer_used = 0.0

    tickets_flat = []

    if not tasks_df.empty:
        planned_sp = tasks_df['sp'].sum()
        planned_tickets = len(tasks_df)
        
        done_df = tasks_df[tasks_df['status'] == 'Done']
        delivered_sp = done_df['sp'].sum()
        completed_tickets = len(done_df)
        spillover_tickets = planned_tickets - completed_tickets

        # Category based buffer metrics
        bug_df = tasks_df[tasks_df['category'] == 'Bug Fix']
        bug_tickets_planned = len(bug_df)
        bug_tickets_done = len(bug_df[bug_df['status'] == 'Done'])
        bug_buffer_used = bug_df['sp'].sum()

        adhoc_df = tasks_df[tasks_df['category'] == 'Adhoc']
        adhoc_buffer_used = adhoc_df['sp'].sum()

        for _, t in tasks_df.iterrows():
            tickets_flat.append({
                "ticket_id": t['ticket_id'],
                "title": t['title'],
                "assignee": t['assignee'],
                "category": t['category'],
                "sp": t['sp'],
                "status": t['status']
            })

    completion_rate_sp = (delivered_sp / planned_sp * 100.0) if planned_sp > 0 else 0.0
    utilization_rate = (delivered_sp / plannable_capacity * 100.0) if plannable_capacity > 0 else 0.0
    bug_resolution_rate = (bug_tickets_done / bug_tickets_planned * 100.0) if bug_tickets_planned > 0 else 0.0
    allocation_rate = (planned_sp / plannable_capacity * 100.0) if plannable_capacity > 0 else 0.0
    avg_team_delivery = sum(d['delivery_rate'] for d in dev_details) / len(dev_details) if dev_details else 0.0
    total_leaves = sum(d.get('leaves', 0) for d in dev_details)
    total_work_days = work_days

    # Summary Generation
    summary = f"Sprint {sprint['name']} was completed with a final delivery of {delivered_sp:.1f} SP out of {planned_sp:.1f} SP committed ({completion_rate_sp:.1f}% completion rate). "
    summary += f"The team had an overall available plannable capacity of {plannable_capacity:.1f} SP, resulting in a {utilization_rate:.1f}% net capacity utilization. "
    summary += f"Out of {planned_tickets} committed tickets, {completed_tickets} were marked Done, leaving {spillover_tickets} as spillover. "
    if bug_tickets_planned > 0:
        summary += f"The team resolved {bug_tickets_done} out of {bug_tickets_planned} production bug fix tickets ({bug_resolution_rate:.1f}% bug resolution rate)."
    else:
        summary += "No production bug fix tickets were committed during this sprint."

    report_data = {
        "sprint_id": str(sprint_id),
        "sprint_name": sprint['name'],
        "team_id": str(team_id),
        "team_name": team_name,
        "planned_start_date": str(sprint['start_date']),
        "planned_end_date": str(sprint['end_date']),
        "actual_start_date": str(sprint.get('actual_start_date', sprint['start_date'])),
        "actual_end_date": str(sprint.get('actual_end_date', sprint['end_date'])),
        "generated_at": datetime.datetime.now().isoformat(),
        "summary": summary,
        "planned_sp": float(planned_sp),
        "delivered_sp": float(delivered_sp),
        "completion_rate_sp": float(completion_rate_sp),
        "total_team_capacity": float(total_net_sp),
        "plannable_capacity": float(plannable_capacity),
        "utilization_rate": float(utilization_rate),
        "allocation_rate": float(allocation_rate),
        "avg_team_delivery": float(avg_team_delivery),
        "planned_tickets": int(planned_tickets),
        "completed_tickets": int(completed_tickets),
        "spillover_tickets": int(spillover_tickets),
        "bug_tickets_planned": int(bug_tickets_planned),
        "bug_tickets_done": int(bug_tickets_done),
        "bug_resolution_rate": float(bug_resolution_rate),
        "bug_buffer_allocated": float(bug_buffer_allocated),
        "bug_buffer_used": float(bug_buffer_used),
        "adhoc_buffer_allocated": float(adhoc_buffer_allocated),
        "adhoc_buffer_used": float(adhoc_buffer_used),
        "ceremony_buffer_allocated": float(ceremony_buffer_allocated),
        "total_leaves": int(total_leaves),
        "total_work_days": int(total_work_days),
        "holiday_count": int(holiday_count),
        "team_size": len(dev_details),
        "dev_details": dev_details,
        "tickets": tickets_flat
    }
    return report_data

def build_pdf_report(report_data):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=54,
        leftMargin=54,
        topMargin=72,
        bottomMargin=72
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        leading=26,
        textColor=colors.HexColor('#1f4e78'),
        spaceAfter=15
    )
    h2_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=13,
        leading=16,
        textColor=colors.HexColor('#1f4e78'),
        spaceBefore=12,
        spaceAfter=6,
        keepWithNext=True
    )
    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9.5,
        leading=13.5,
        textColor=colors.HexColor('#334155'),
        spaceAfter=8
    )
    table_header_style = ParagraphStyle(
        'TableHeaderCustom',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=10.5,
        textColor=colors.white
    )
    table_cell_style = ParagraphStyle(
        'TableCellCustom',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=10.5,
        textColor=colors.HexColor('#334155')
    )
    table_cell_bold = ParagraphStyle(
        'TableCellBold',
        parent=table_cell_style,
        fontName='Helvetica-Bold'
    )
    
    elements = []
    
    # Title
    elements.append(Paragraph(f"Sprint Performance Report: {report_data['sprint_name']}", title_style))
    elements.append(Paragraph(f"<b>Team:</b> {report_data['team_name']} &nbsp;|&nbsp; <b>Planned Dates:</b> {report_data['planned_start_date']} to {report_data['planned_end_date']} &nbsp;|&nbsp; <b>Actual:</b> {report_data['actual_start_date']} to {report_data['actual_end_date']}", body_style))
    elements.append(Spacer(1, 10))
    
    # Summary Section
    elements.append(Paragraph("Executive Summary", h2_style))
    elements.append(Paragraph(report_data['summary'], body_style))
    elements.append(Spacer(1, 10))
    
    # KPI Grid Table
    elements.append(Paragraph("Key Performance Metrics", h2_style))
    kpi_data = [
        [
            Paragraph("Metric", table_header_style), 
            Paragraph("Planned / Target", table_header_style), 
            Paragraph("Actual / Delivered", table_header_style),
            Paragraph("Variance / Status", table_header_style)
        ],
        [Paragraph("Committed Story Points (SP)", table_cell_style), Paragraph(f"{report_data['planned_sp']:.1f}", table_cell_style), Paragraph(f"{report_data['delivered_sp']:.1f}", table_cell_style), Paragraph(f"{report_data['delivered_sp'] - report_data['planned_sp']:.1f} ({report_data['completion_rate_sp']:.1f}%)", table_cell_style)],
        [Paragraph("Capacity (Plannable)", table_cell_style), Paragraph(f"{report_data['plannable_capacity']:.1f}", table_cell_style), Paragraph(f"{report_data['delivered_sp']:.1f}", table_cell_style), Paragraph(f"Utilization: {report_data['utilization_rate']:.1f}%", table_cell_style)],
        [Paragraph("Total Commits (Tickets)", table_cell_style), Paragraph(f"{report_data['planned_tickets']}", table_cell_style), Paragraph(f"{report_data['completed_tickets']}", table_cell_style), Paragraph(f"Spillover: {report_data['spillover_tickets']}", table_cell_style)],
        [Paragraph("Production Bug Fixes", table_cell_style), Paragraph(f"{report_data['bug_tickets_planned']}", table_cell_style), Paragraph(f"{report_data['bug_tickets_done']}", table_cell_style), Paragraph(f"Resolution Rate: {report_data['bug_resolution_rate']:.1f}%", table_cell_style)],
    ]
    t_kpis = Table(kpi_data, colWidths=[150, 110, 110, 134])
    t_kpis.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1f4e78')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 4),
        ('TOPPADDING', (0,1), (-1,-1), 4),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
    ]))
    elements.append(t_kpis)
    elements.append(Spacer(1, 12))
    
    # Buffer Allocation & Usage
    elements.append(Paragraph("Buffer Utilization & Consumption", h2_style))
    buffer_data = [
        [Paragraph("Buffer Type", table_header_style), Paragraph("Allocated (SP)", table_header_style), Paragraph("Actual Usage (SP)", table_header_style), Paragraph("Usage Status", table_header_style)],
        [Paragraph("Production Bug Buffer", table_cell_style), Paragraph(f"{report_data['bug_buffer_allocated']:.1f}", table_cell_style), Paragraph(f"{report_data['bug_buffer_used']:.1f}", table_cell_style), Paragraph("Within Limits" if report_data['bug_buffer_used'] <= report_data['bug_buffer_allocated'] else "Exceeded", table_cell_bold if report_data['bug_buffer_used'] > report_data['bug_buffer_allocated'] else table_cell_style)],
        [Paragraph("Adhoc Buffer", table_cell_style), Paragraph(f"{report_data['adhoc_buffer_allocated']:.1f}", table_cell_style), Paragraph(f"{report_data['adhoc_buffer_used']:.1f}", table_cell_style), Paragraph("Within Limits" if report_data['adhoc_buffer_used'] <= report_data['adhoc_buffer_allocated'] else "Exceeded", table_cell_bold if report_data['adhoc_buffer_used'] > report_data['adhoc_buffer_allocated'] else table_cell_style)],
        [Paragraph("Ceremony & Administrative", table_cell_style), Paragraph(f"{report_data['ceremony_buffer_allocated']:.1f}", table_cell_style), Paragraph("N/A", table_cell_style), Paragraph("Full usage assumed", table_cell_style)],
    ]
    t_buffer = Table(buffer_data, colWidths=[150, 110, 110, 134])
    t_buffer.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1f4e78')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 4),
        ('TOPPADDING', (0,1), (-1,-1), 4),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
    ]))
    elements.append(t_buffer)
    elements.append(Spacer(1, 12))
    
    # Team Roster & Delivery breakdown
    elements.append(Paragraph("Developer Wise Workload & Delivery Breakdown", h2_style))
    dev_header = [
        Paragraph("Resource Name", table_header_style), 
        Paragraph("Role", table_header_style), 
        Paragraph("Net Capacity", table_header_style),
        Paragraph("Committed", table_header_style),
        Paragraph("Completed", table_header_style),
        Paragraph("Delivery %", table_header_style)
    ]
    dev_table_rows = [dev_header]
    for dev in report_data['dev_details']:
        dev_table_rows.append([
            Paragraph(dev['name'], table_cell_bold),
            Paragraph(dev['role'], table_cell_style),
            Paragraph(f"{dev['capacity']:.1f}", table_cell_style),
            Paragraph(f"{dev['allocated']:.1f}", table_cell_style),
            Paragraph(f"{dev['completed']:.1f}", table_cell_style),
            Paragraph(f"{dev['delivery_rate']:.1f}%", table_cell_style),
        ])
    t_dev = Table(dev_table_rows, colWidths=[120, 70, 70, 70, 70, 104])
    t_dev.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1f4e78')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 4),
        ('TOPPADDING', (0,1), (-1,-1), 4),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
    ]))
    elements.append(t_dev)
    
    # Page Break for Detail Backlog (only if tickets are present)
    if report_data['tickets']:
        elements.append(PageBreak())
        elements.append(Paragraph("Committed Backlog Detail List", h2_style))
        elements.append(Spacer(1, 5))
        backlog_header = [
            Paragraph("Ticket", table_header_style),
            Paragraph("Title", table_header_style),
            Paragraph("Assignee", table_header_style),
            Paragraph("Category", table_header_style),
            Paragraph("SP", table_header_style),
            Paragraph("Status", table_header_style)
        ]
        backlog_rows = [backlog_header]
        for ticket in report_data['tickets']:
            backlog_rows.append([
                Paragraph(ticket['ticket_id'], table_cell_bold),
                Paragraph(ticket['title'][:40] + ("..." if len(ticket['title']) > 40 else ""), table_cell_style),
                Paragraph(ticket['assignee'], table_cell_style),
                Paragraph(ticket['category'], table_cell_style),
                Paragraph(f"{ticket['sp']:.1f}", table_cell_style),
                Paragraph(ticket['status'], table_cell_style),
            ])
        t_backlog = Table(backlog_rows, colWidths=[64, 180, 80, 70, 40, 70])
        t_backlog.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1f4e78')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('BOTTOMPADDING', (0,0), (-1,0), 5),
            ('TOPPADDING', (0,0), (-1,0), 5),
            ('BOTTOMPADDING', (0,1), (-1,-1), 4),
            ('TOPPADDING', (0,1), (-1,-1), 4),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ]))
        elements.append(t_backlog)
    
    doc.build(elements, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer.getvalue()

def build_excel_report(report_data):
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=10)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # 1. Summary Sheet
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1["A1"] = f"Sprint Performance Summary: {report_data['sprint_name']}"
    ws1["A1"].font = title_font
    ws1.row_dimensions[1].height = 30
    
    ws1["A3"] = f"Team: {report_data['team_name']}"
    ws1["A3"].font = bold_font
    ws1["A4"] = f"Planned Dates: {report_data['planned_start_date']} to {report_data['planned_end_date']}"
    ws1["A4"].font = regular_font
    ws1["A5"] = f"Actual Dates: {report_data['actual_start_date']} to {report_data['actual_end_date']}"
    ws1["A5"].font = regular_font
    
    summary_headers = ["KPI Metric", "Target / Planned", "Actual / Delivered", "Completion / Status"]
    for col_idx, text in enumerate(summary_headers, start=1):
        cell = ws1.cell(row=7, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[7].height = 24
    
    summary_rows = [
        ["Committed Story Points", report_data['planned_sp'], report_data['delivered_sp'], f"{report_data['completion_rate_sp']:.1f}% Completion"],
        ["Capacity (Plannable)", report_data['plannable_capacity'], report_data['delivered_sp'], f"{report_data['utilization_rate']:.1f}% Net Utilization"],
        ["Total Tickets", report_data['planned_tickets'], report_data['completed_tickets'], f"{report_data['spillover_tickets']} Spillover"],
        ["Production Bug Resolution", report_data['bug_tickets_planned'], report_data['bug_tickets_done'], f"{report_data['bug_resolution_rate']:.1f}% Resolved"],
        ["Bug Buffer Usage (SP)", report_data['bug_buffer_allocated'], report_data['bug_buffer_used'], "Within Limits" if report_data['bug_buffer_used'] <= report_data['bug_buffer_allocated'] else "Exceeded"],
        ["Adhoc Buffer Usage (SP)", report_data['adhoc_buffer_allocated'], report_data['adhoc_buffer_used'], "Within Limits" if report_data['adhoc_buffer_used'] <= report_data['adhoc_buffer_allocated'] else "Exceeded"],
        ["Ceremony Buffer (SP)", report_data['ceremony_buffer_allocated'], "N/A", "Consumed"],
    ]
    
    for row_idx, r in enumerate(summary_rows, start=8):
        for col_idx, val in enumerate(r, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
            if col_idx in [2, 3] and isinstance(val, (int, float)):
                cell.number_format = '0.0'
                cell.alignment = Alignment(horizontal="right")
        ws1.row_dimensions[row_idx].height = 20
        
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # 2. Roster Sheet
    ws2 = wb.create_sheet(title="Developer Capacity")
    ws2.views.sheetView[0].showGridLines = True
    ws2["A1"] = "Developer-wise Performance Breakdown"
    ws2["A1"].font = title_font
    
    roster_headers = ["Resource Name", "Role", "Net Capacity", "Committed SP", "Completed SP", "Delivery Rate"]
    for col_idx, text in enumerate(roster_headers, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[3].height = 24
    
    for row_idx, dev in enumerate(report_data['dev_details'], start=4):
        ws2.cell(row=row_idx, column=1, value=dev['name']).font = bold_font
        ws2.cell(row=row_idx, column=2, value=dev['role']).font = regular_font
        ws2.cell(row=row_idx, column=3, value=dev['capacity']).number_format = '0.0'
        ws2.cell(row=row_idx, column=4, value=dev['allocated']).number_format = '0.0'
        ws2.cell(row=row_idx, column=5, value=dev['completed']).number_format = '0.0'
        
        del_rate_cell = ws2.cell(row=row_idx, column=6, value=f"{dev['delivery_rate']:.1f}%")
        del_rate_cell.font = regular_font
        
        for c in range(1, 7):
            cell = ws2.cell(row=row_idx, column=c)
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
            if c in [3, 4, 5]:
                cell.alignment = Alignment(horizontal="right")
        ws2.row_dimensions[row_idx].height = 20
        
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # 3. Backlog Sheet
    ws3 = wb.create_sheet(title="Backlog Detail")
    ws3.views.sheetView[0].showGridLines = True
    ws3["A1"] = "Committed Backlog Detail"
    ws3["A1"].font = title_font
    
    backlog_headers = ["Ticket ID", "Title", "Assignee", "Category", "Story Points", "Status"]
    for col_idx, text in enumerate(backlog_headers, start=1):
        cell = ws3.cell(row=3, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[3].height = 24
    
    for row_idx, ticket in enumerate(report_data['tickets'], start=4):
        ws3.cell(row=row_idx, column=1, value=ticket['ticket_id']).font = bold_font
        ws3.cell(row=row_idx, column=2, value=ticket['title']).font = regular_font
        ws3.cell(row=row_idx, column=3, value=ticket['assignee']).font = regular_font
        ws3.cell(row=row_idx, column=4, value=ticket['category']).font = regular_font
        
        sp_cell = ws3.cell(row=row_idx, column=5, value=ticket['sp'])
        sp_cell.number_format = '0.0'
        sp_cell.font = regular_font
        
        ws3.cell(row=row_idx, column=6, value=ticket['status']).font = regular_font
        
        for c in range(1, 7):
            cell = ws3.cell(row=row_idx, column=c)
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
            if c == 5:
                cell.alignment = Alignment(horizontal="right")
        ws3.row_dimensions[row_idx].height = 20
        
    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def build_png_report(report_data):
    img = PILImage.new('RGB', (1200, 900), color='#0f172a')
    draw = ImageDraw.Draw(img)

    try:
        font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        font_path_regular = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        title_font = ImageFont.truetype(font_path, 28)
        sub_font = ImageFont.truetype(font_path, 14)
        kpi_val_font = ImageFont.truetype(font_path, 32)
        kpi_lbl_font = ImageFont.truetype(font_path, 12)
        body_font = ImageFont.truetype(font_path_regular, 13)
        body_bold = ImageFont.truetype(font_path, 13)
        small_font = ImageFont.truetype(font_path_regular, 11)
    except IOError:
        title_font = ImageFont.load_default()
        sub_font = ImageFont.load_default()
        kpi_val_font = ImageFont.load_default()
        kpi_lbl_font = ImageFont.load_default()
        body_font = ImageFont.load_default()
        body_bold = ImageFont.load_default()
        small_font = ImageFont.load_default()

    # Header
    draw.rectangle([0, 0, 1200, 100], fill="#1e293b", outline="#334155", width=1)
    draw.text((40, 15), f"Sprint Performance Dashboard", fill="#38bdf8", font=title_font)
    draw.text((40, 55), f"Sprint: {report_data['sprint_name']}  |  Team: {report_data['team_name']}  |  {report_data['planned_start_date']} to {report_data['actual_end_date']}", fill="#94a3b8", font=small_font)
    draw.text((40, 75), f"Generated: {report_data.get('generated_at', '')[:19].replace('T', ' ')}", fill="#64748b", font=small_font)

    # Primary KPIs row
    col_width = 285
    gap = 15
    kpis = [
        {"val": f"{report_data['planned_sp']:.1f}", "lbl": "Committed SP", "color": "#a78bfa"},
        {"val": f"{report_data['delivered_sp']:.1f}", "lbl": "Delivered SP", "color": "#4ade80"},
        {"val": f"{report_data['completion_rate_sp']:.0f}%", "lbl": "Completion Rate", "color": "#60a5fa"},
        {"val": f"{report_data['utilization_rate']:.0f}%", "lbl": "Net Utilization", "color": "#f472b6"}
    ]

    for idx, kpi in enumerate(kpis):
        x1 = 30 + idx * (col_width + gap)
        y1 = 115
        x2 = x1 + col_width
        y2 = 205
        draw.rectangle([x1, y1, x2, y2], fill="#1e293b", outline="#334155", width=1)
        draw.text((x1 + 15, y1 + 15), kpi["val"], fill=kpi["color"], font=kpi_val_font)
        draw.text((x1 + 15, y1 + 60), kpi["lbl"], fill="#94a3b8", font=kpi_lbl_font)

    # Left panel - Summary & Team
    draw.rectangle([30, 225, 620, 580], fill="#1e293b", outline="#334155", width=1)
    draw.text((50, 240), "Executive Summary", fill="#38bdf8", font=sub_font)

    summary_text = report_data['summary']
    words = summary_text.split()
    lines = []
    current_line = []
    for w in words:
        current_line.append(w)
        if len(" ".join(current_line)) > 60:
            lines.append(" ".join(current_line[:-1]))
            current_line = [w]
    if current_line:
        lines.append(" ".join(current_line))

    y_offset = 265
    for l in lines[:6]:
        draw.text((50, y_offset), l, fill="#cbd5e1", font=body_font)
        y_offset += 18

    draw.text((50, y_offset + 10), "Team Performance", fill="#38bdf8", font=sub_font)
    table_y = y_offset + 35
    draw.text((50, table_y), "Name", fill="#94a3b8", font=body_bold)
    draw.text((180, table_y), "Role", fill="#94a3b8", font=body_bold)
    draw.text((270, table_y), "Capacity", fill="#94a3b8", font=body_bold)
    draw.text((360, table_y), "Allocated", fill="#94a3b8", font=body_bold)
    draw.text((450, table_y), "Done %", fill="#94a3b8", font=body_bold)

    for dev in report_data['dev_details'][:5]:
        table_y += 22
        draw.text((50, table_y), dev['name'][:15], fill="#e2e8f0", font=body_font)
        draw.text((180, table_y), dev['role'][:8], fill="#cbd5e1", font=body_font)
        draw.text((270, table_y), f"{dev['capacity']:.1f}", fill="#cbd5e1", font=body_font)
        draw.text((360, table_y), f"{dev['allocated']:.1f}", fill="#cbd5e1", font=body_font)
        draw.text((450, table_y), f"{dev['delivery_rate']:.0f}%", fill="#4ade80" if dev['delivery_rate'] >= 80 else "#fb7185", font=body_font)

    # Right panel - Buffer & Metrics
    draw.rectangle([640, 225, 1170, 580], fill="#1e293b", outline="#334155", width=1)
    draw.text((660, 240), "Capacity & Buffers", fill="#38bdf8", font=sub_font)

    metrics = [
        ("Total Team Capacity", f"{report_data['total_team_capacity']:.1f} SP"),
        ("Plannable Capacity", f"{report_data['plannable_capacity']:.1f} SP"),
        ("Allocation Rate", f"{report_data.get('allocation_rate', 0):.1f}%"),
        ("Bug Buffer", f"{report_data['bug_buffer_used']:.1f} / {report_data['bug_buffer_allocated']:.1f} SP"),
        ("Adhoc Buffer", f"{report_data['adhoc_buffer_used']:.1f} / {report_data['adhoc_buffer_allocated']:.1f} SP"),
        ("Ceremony Buffer", f"{report_data['ceremony_buffer_allocated']:.1f} SP"),
    ]

    metric_y = 265
    for lbl, val in metrics:
        draw.text((660, metric_y), lbl, fill="#94a3b8", font=body_font)
        draw.text((1050, metric_y), val, fill="#e2e8f0", font=body_bold)
        metric_y += 22

    # Buffer progress bars
    bar_y = metric_y + 20
    bug_p = (report_data['bug_buffer_used'] / report_data['bug_buffer_allocated']) if report_data['bug_buffer_allocated'] > 0 else 0
    draw.text((660, bar_y), "Bug Buffer Usage", fill="#94a3b8", font=small_font)
    draw.rectangle([660, bar_y + 18, 1140, bar_y + 30], fill="#0f172a", outline="#334155")
    draw.rectangle([660, bar_y + 18, 660 + int(480 * min(bug_p, 1)), bar_y + 30], fill="#f87171")

    bar_y += 45
    adhoc_p = (report_data['adhoc_buffer_used'] / report_data['adhoc_buffer_allocated']) if report_data['adhoc_buffer_allocated'] > 0 else 0
    draw.text((660, bar_y), "Adhoc Buffer Usage", fill="#94a3b8", font=small_font)
    draw.rectangle([660, bar_y + 18, 1140, bar_y + 30], fill="#0f172a", outline="#334155")
    draw.rectangle([660, bar_y + 18, 660 + int(480 * min(adhoc_p, 1)), bar_y + 30], fill="#fb923c")

    # Bottom panel - Sprint Activity
    draw.rectangle([30, 600, 1170, 880], fill="#1e293b", outline="#334155", width=1)
    draw.text((50, 615), "Sprint Activity & Resolution Highlights", fill="#38bdf8", font=sub_font)

    stat_rows = [
        ("Total Tickets Committed", f"{report_data['planned_tickets']}"),
        ("Completed & Delivered", f"{report_data['completed_tickets']}"),
        ("Spillover/Incomplete", f"{report_data['spillover_tickets']}"),
        ("Production Bugs Committed", f"{report_data['bug_tickets_planned']}"),
        ("Production Bugs Resolved", f"{report_data['bug_tickets_done']}"),
        ("Bug Resolution Rate", f"{report_data['bug_resolution_rate']:.1f}%"),
        ("Team Size", f"{report_data.get('team_size', 'N/A')}"),
        ("Total Work Days", f"{report_data.get('total_work_days', 'N/A')}"),
    ]

    col1_y = 645
    col2_y = 645
    for i, (lbl, val) in enumerate(stat_rows):
        if i < 4:
            draw.text((50, col1_y), lbl, fill="#94a3b8", font=body_font)
            draw.text((280, col1_y), val, fill="#e2e8f0", font=body_bold)
            col1_y += 28
        else:
            draw.text((400, col2_y), lbl, fill="#94a3b8", font=body_font)
            draw.text((620, col2_y), val, fill="#e2e8f0", font=body_bold)
            col2_y += 28

    # Footer
    draw.text((50, 855), "Confidential - For Internal Stakeholder Review", fill="#64748b", font=small_font)

    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    return buffer.getvalue()

def compile_and_save_report(sprint_id):
    """Compiles all sprint stats, generates binary exports, and persists in MongoDB."""
    report_data = generate_sprint_report_data(sprint_id)
    
    # Generate export files
    pdf_bytes = build_pdf_report(report_data)
    excel_bytes = build_excel_report(report_data)
    png_bytes = build_png_report(report_data)
    
    # Add binary files to document
    report_doc = report_data.copy()
    report_doc['pdf_data'] = bson.Binary(pdf_bytes)
    report_doc['excel_data'] = bson.Binary(excel_bytes)
    report_doc['png_data'] = bson.Binary(png_bytes)
    
    from utils.db import save_sprint_report
    save_sprint_report(report_doc)
    return report_doc
